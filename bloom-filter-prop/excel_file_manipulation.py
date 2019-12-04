import time
import os.path
import glob
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import xlrd


def write_results_to_excel_file(spec, runs, output, process_id, path_name):
    """
    This function generates the excel files from the data created.

    When working in a multiprocessing mode this will produce one excel spreadsheet per process that is running.
    This is done to avoid deadlock with multiple processes opening the file at once
    One sheet is created per P value
    """

    wb = Workbook()
    if os.path.isfile(path_name):
        wb = load_workbook(path_name)

    sheet_result_title = "P_" + str(spec['num_of_producers'])

    if sheet_result_title not in wb.sheetnames:
        wb.create_sheet(sheet_result_title)

    sheet_result = initiate_worksheet(wb, sheet_result_title)
    ind_col = 1
    ind_row = 1
    while sheet_result.cell(row=ind_row, column=ind_col).value is not None:
        ind_row += 1
    # parameters
    for col_exc in spec:
        sheet_result.cell(row=ind_row, column=ind_col).value = spec[col_exc]
        ind_col += 1
    # runs
    sheet_result.cell(row=ind_row, column=ind_col).value = runs
    ind_col += 1
    # output
    for col_out in output:
        sheet_result.cell(row=ind_row, column=ind_col).value = output[col_out]
        ind_col += 1
    # processID
    sheet_result.cell(row=ind_row, column=ind_col).value = process_id
    ind_col += 1

    wb.save(path_name)


def initiate_worksheet(workbook, sheet_title):
    """
    This function generates the the indevidual worksheets in the book being created by write_results_to_excel_file
    It also creates the collumn titles used
    """
    sheet_result = workbook.worksheets[workbook.index(workbook[sheet_title])]
    ind_col = 1
    if sheet_result['A1'].value is None:
        name_cols = ['Total Producers', 'Correct Producers Ratio', 'Collected Updates Ratio', 'Collected Votes Ratio',
                     'Collected Final Votes Ratio', 'runs', 'Total Correct Ln(prod)',
                     'Total Correct Ln(vote)', 'Runs With All Ln(prod)',
                     'Runs With All Ln(vote)',
                     'Runs With > 50% Correct', 'Runs With = Cn']
        for name_col in name_cols:
            sheet_result.cell(row=1, column=ind_col).value = name_col
            sheet_result.column_dimensions[get_column_letter(ind_col)].width = 20
            sheet_result.cell(row=1, column=ind_col).alignment = Alignment(wrap_text=True)
            ind_col += 1
    return sheet_result


def combine_excel_files(end_producer, step_producer, spec):
    """
    This function combines all the indevidual excel files created through multiprocessing into a master file
    """
    glob.glob("excel/*.xlsx")
    timestr = get_time()
    start_producer = spec['num_of_producers']
    try:
        if not os.listdir('merged-excel-docs'):
            print('Folder empty no need to remove files')
            os.mkdir('merged-excel-docs')
    except FileNotFoundError:
        os.mkdir('merged-excel-docs')

    writer = pd.ExcelWriter('merged-excel-docs/combined-result' + timestr + '.xlsx', engine='xlsxwriter')
    for ind_p in range(start_producer, end_producer, step_producer):
        all_data = pd.DataFrame()
        sheetID = str(ind_p)
        for f in glob.glob("excel/*.xlsx"):
            df = pd.read_excel(f, "P_" + sheetID)
            all_data = all_data.append(df, ignore_index=True)
        all_data.to_excel(writer, sheet_name="P_" + sheetID)
    writer.save()


def move_old_excel():
    """
    This function is used to move all old exel files into a cache to store for later use without affecting the current run of the script.

    These old files can still be accessed and are used by combine_global_output_file
    """
    timestr = get_time()

    try:
        if not os.listdir('old_excel'):
            print('Folder empty no need to remove files')
    except FileNotFoundError:
        os.mkdir('old_excel')

    print("passing here")
    try:
        if not os.listdir('excel'):
            print('Folder empty no need to remove files')
        else:
            os.rename('excel', 'old_excel/excel_' + timestr)
            os.mkdir('excel')
            print("created folder")
    except FileNotFoundError:
        os.mkdir('excel')
        print("created folder within exception")


def get_time():
    """
    Get the time stamp for file names
    """
    return time.strftime("%Y%m%d-%H%M%S")


def combine_global_output_file(end_producer, step_producer, spec):
    """
    This function creates a global file. This takes all data that has been stored from previous runs of the script.
    It then generates the global file
    Currently end_producer value needs to be changed in order to show the max P_XXX value we have in any of the files

    NOTE: IF MAJOR CHANGES ARE PERFORMED ON THE CODE THEN THE CACHE OF PREVIOUS EXCEL FILES MUST BE CLEARED.
    """

    start_producer = 100
    end_producer = 1001
    writer = pd.ExcelWriter('global_output.xlsx', engine='xlsxwriter')

    for ind_p in range(start_producer, end_producer, step_producer):
        all_data = pd.DataFrame()
        sheetID = str(ind_p)
        for f in glob.glob("merged-excel-docs/*.xlsx"):
            try:
                df = pd.read_excel(f, "P_" + sheetID)
                all_data = all_data.append(df, ignore_index=True)
            except (ValueError, xlrd.biffh.XLRDError):
                continue
        all_data.to_excel(writer, sheet_name="P_" + sheetID)
    writer.save()


def generate_postprocessed_files():
    """
    This function takes the global file and is used to form a processed file in which any duplicated of runs are merged together.
    It allows us to see the data accross many runs, thereby conglomerating the data.

    NOTE: IF MAJOR CHANGES ARE PERFORMED ON THE CODE THEN THE CACHE OF PREVIOUS EXCEL FILES MUST BE CLEARED.

    NOTE: this file now splits to the varying P_XXX values as with all other files. The master file generated now is good to go.
    """
    get_excel_file = pd.ExcelFile('global_output.xlsx')
    get_sheet_names = get_excel_file.sheet_names

    writer = pd.ExcelWriter('master_ouput.xlsx', engine='xlsxwriter')
    for sheet in get_sheet_names:
        try:
            all_data = pd.DataFrame()
            sheetID = str(sheet)
            data = pd.read_excel('global_output.xlsx', sheet, dtype={'id': str})
            grouped_data = data.groupby(['Total Producers', 'Correct Producers Ratio', 'Collected Updates Ratio',
                                         'Collected Votes Ratio', 'Collected Final Votes Ratio'], as_index=False)[
                'Total Correct Ln(prod)',
                'runs', 'Total Correct Ln(vote)',
                'Runs With All Ln(prod)',
                'Runs With All Ln(vote)',
                'Runs With > 50% Correct', 'Runs With = Cn'].sum()

            grouped_data['num_correct_producers_Ln_prod'] = grouped_data['Total Correct Ln(prod)'] / grouped_data[
                'runs']
            grouped_data['num_correct_producers_Ln_vote'] = grouped_data['Total Correct Ln(vote)'] / grouped_data[
                'runs']
            grouped_data['percentage_for_50_%'] = (grouped_data['Runs With > 50% Correct'] / grouped_data['runs']) * 100
            grouped_data['Percentage Runs With = Cn'] = (grouped_data['Runs With = Cn'] / grouped_data['runs']) * 100

            all_data = all_data.append(grouped_data, ignore_index=True)

            all_data.to_excel(writer, sheet_name=sheet)
        except KeyError:
            continue
    writer.save()
    print("Merged File")